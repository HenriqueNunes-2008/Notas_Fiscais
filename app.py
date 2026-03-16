from flask import Flask, render_template, request, jsonify
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__)

# Caminho para a planilha na mesma pasta do script
EXCEL_PATH = "nota_almoxarifado.xlsx"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/adicionar', methods=['POST'])
def adicionar():
    try:
        # 1. Coleta os dados do formulário
        data_nota = request.form.get('data')
        n_nota = request.form.get('n_nota')
        fornecedor = request.form.get('fornecedor')
        materiais = request.form.getlist('material[]')
        quantidades = request.form.getlist('quantidade[]')
        valores_unit = request.form.getlist('valor_unitario[]')
        unidades = request.form.getlist('unidade[]')
        projetos = request.form.getlist('projeto[]')

        if not materiais:
            return jsonify({"status": "error", "message": "Nenhum material informado."})

        # 2. Prepara as linhas para salvar
        novas_linhas = []
        for i in range(len(materiais)):
            qtd = float(quantidades[i] or 0)
            v_unit = float(valores_unit[i] or 0)
            v_total = qtd * v_unit
            
            # Ordem exata das colunas:
            novas_linhas.append([
                data_nota, n_nota, fornecedor, materiais[i], 
                unidades[i], qtd, v_unit, v_total, projetos[i]
            ])

        # 3. Lógica de Salvamento na Planilha
        colunas = ["Data", "N nota", "Fornecedor", "Material", "Unidade", "Quantidade", "Valor unitario", "Valor Total", "Projeto"]

        if not os.path.exists(EXCEL_PATH):
            # Cria a planilha do zero se não existir
            df = pd.DataFrame(novas_linhas, columns=colunas)
            df.to_excel(EXCEL_PATH, sheet_name='dados', index=False)
        else:
            # Abre a planilha existente e adiciona na aba 'dados'
            book = load_workbook(EXCEL_PATH)
            if 'dados' not in book.sheetnames:
                book.create_sheet('dados')
                sheet = book['dados']
                sheet.append(colunas) # Adiciona cabeçalho se a aba for nova
            else:
                sheet = book['dados']

            for linha in novas_linhas:
                sheet.append(linha)
            
            book.save(EXCEL_PATH)
            book.close()

        return jsonify({
            "status": "success", 
            "message": f"✅ {len(materiais)} itens salvos com sucesso na planilha!"
        })

    except PermissionError:
        return jsonify({"status": "error", "message": "⚠️ O Excel está aberto! Feche o arquivo e tente novamente."})
    except Exception as e:
        return jsonify({"status": "error", "message": f"❌ Erro: {str(e)}"})

if __name__ == '__main__':
    app.run(debug=True)